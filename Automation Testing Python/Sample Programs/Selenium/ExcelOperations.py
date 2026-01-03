import openpyxl
filename = "c:\\UserIDPass.xlsx"
# Load/ open excel file
workbook = openpyxl.load_workbook(filename)

# Locate/get perticular sheet
sheet = workbook["Sheet1"]                  # workbook.get_sheet_by_name("Sheet1")

# Read data from sheet
print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=3, column=2).value)

temp = sheet.cell(row=1, column=1).value

sheet.cell(row=6, column=1).value = "Sitaram"
sheet.cell(row=6, column=2).value = "SitaramIsKing"

print(sheet.cell(row=6, column=1).value)
print(sheet.cell(row=6, column=2).value)

workbook.save(filename)